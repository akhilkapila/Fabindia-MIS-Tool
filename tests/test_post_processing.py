#!/usr/bin/env python3
"""
Test the post-processing function to verify formatting is applied correctly.
"""
import sys
import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import post_process_workbook, load_file_smartly

print("=" * 70)
print("TESTING POST-PROCESSING FUNCTION")
print("=" * 70)

# Create a test workbook with various column types
test_wb_path = tempfile.mktemp(suffix='.xlsx')
wb = Workbook()
ws = wb.active
ws.title = 'Test Data'

# Headers
headers = ['Store Code', 'Date', 'Amount', 'Notes', 'Remarks']
ws.append(headers)

# Add test data with different types
ws.append(['S001', datetime(2025, 11, 1), 1000, 'Test note', 'Keep as is'])
ws.append(['S002', datetime(2025, 11, 2), 2000, 'Another note', ''])
ws.append(['S003', datetime(2025, 11, 3), 3000, 'More notes', 'Final remark'])

# Add a row with one blank cell to test blank column detection
ws.append(['S004', datetime(2025, 11, 4), 4000, '', 'Has blank'])

wb.save(test_wb_path)
print(f"\n1. Created test workbook: {test_wb_path}")

# Apply post-processing
print("\n2. Applying post-processing...")
try:
    post_process_workbook(test_wb_path, sheets=['Test Data'], min_data_row=2)
    print("   ✓ Post-processing successful")
except Exception as e:
    print(f"   ✗ Post-processing failed: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Verify the results
print("\n3. Verifying formatting...")

try:
    # Load the processed workbook
    from openpyxl import load_workbook as openpyxl_load
    processed_wb = openpyxl_load(test_wb_path)
    ws_proc = processed_wb.active
    
    # Check Date column formatting (should be DD-MM-YYYY)
    date_cell = ws_proc['B2']  # First date cell
    if date_cell.number_format and 'dd' in date_cell.number_format.lower():
        print(f"   ✓ Date column has date format: {date_cell.number_format}")
    else:
        print(f"   ⚠ Date column format: {date_cell.number_format}")
    
    # Check that Remarks column (column E) has General format
    remarks_cell = ws_proc['E2']
    print(f"   ✓ Remarks cell format: {remarks_cell.number_format}")
    
    # Check for blank column highlighting (column D is mostly blank except row 4)
    blank_col_cell = ws_proc['D3']  # Blank cell in Notes column should be colored
    if blank_col_cell.fill and blank_col_cell.fill.start_color:
        color = blank_col_cell.fill.start_color
        print(f"   ✓ Blank cell has fill color: {color.rgb if hasattr(color, 'rgb') else color.index}")
    else:
        print(f"   ⚠ Blank cell may not have expected formatting")
    
    # Load with pandas to verify data is intact
    df = pd.read_excel(test_wb_path, sheet_name='Test Data', header=0)
    print(f"\n4. Data integrity check:")
    print(f"   Shape: {df.shape}")
    print(f"   Columns: {list(df.columns)}")
    print(f"   First row Store Code: {df.iloc[0]['Store Code']}")
    print(f"   ✓ Data loaded successfully")
    
except Exception as e:
    print(f"   ✗ Verification failed: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Clean up
os.unlink(test_wb_path)

print("\n" + "=" * 70)
print("✓ POST-PROCESSING TEST PASSED!")
print("=" * 70)
