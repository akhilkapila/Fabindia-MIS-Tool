import os
import io
import urllib.request
import urllib.error
import uuid
from openpyxl import Workbook

BASE_URL = 'http://127.0.0.1:5001'
OUT_DIR = 'temp_uploads'
os.makedirs(OUT_DIR, exist_ok=True)

# Helper to create an Excel with two blank rows then header row at row 3
def create_mis_working(path, rows=3):
    wb = Workbook()
    ws = wb.active
    ws.title = 'MIS Working'
    # Add two blank rows
    ws.append([])
    ws.append([])
    # Header row
    headers = ['Store Code', 'Date', 'HB-Card', 'HB-Cash']
    ws.append(headers)
    # Add some data rows
    data = [
        ['S001', '2025-11-01', 100, 50],
        ['S002', '2025-11-02', 200, 80],
    ]
    for r in data:
        ws.append(r)
    wb.save(path)

# Helper to create Final MIS with Reconciliation by Date by Store sheet
def create_final_mis(path):
    wb = Workbook()
    # First sheet is Reconciliation by Date by Store
    ws = wb.active
    ws.title = 'Reconciliation by Date by Store'
    # Add header row at row 1 (function will try to detect by heuristics)
    headers = ['Store Code', 'Date', 'HB-Card', 'HB-Cash', 'OtherSheetCol']
    ws.append(headers)
    data = [
        ['S001', '2025-11-01', None, None, 'x'],
        ['S002', '2025-11-02', None, None, 'y'],
    ]
    for r in data:
        ws.append(r)
    # Add another sheet to ensure it's preserved
    wb.create_sheet('Summary')
    wb.save(path)

if __name__ == '__main__':
    combine_path = os.path.join(OUT_DIR, 'sample_combine.xlsx')
    final_path = os.path.join(OUT_DIR, 'sample_final.xlsx')
    create_mis_working(combine_path)
    create_final_mis(final_path)
    print('Created sample files:')
    print(' -', combine_path)
    print(' -', final_path)

    # Helper: multipart POST using urllib (so no third-party requests needed)
    def post_multipart(url, files):
        boundary = '----WebKitFormBoundary' + uuid.uuid4().hex
        data_parts = []
        for fieldname, filename, filebytes, mimetype in files:
            data_parts.append(f'--{boundary}')
            data_parts.append(f'Content-Disposition: form-data; name="{fieldname}"; filename="{filename}"')
            data_parts.append(f'Content-Type: {mimetype}')
            data_parts.append('')
            data_parts.append(filebytes)
        data_parts.append(f'--{boundary}--')
        body = b"\r\n".join([part if isinstance(part, bytes) else part.encode('utf-8') for part in data_parts])
        req = urllib.request.Request(url, data=body)
        req.add_header('Content-Type', f'multipart/form-data; boundary={boundary}')
        req.add_header('Content-Length', str(len(body)))
        try:
            with urllib.request.urlopen(req) as resp:
                return resp.status, resp.read()
        except urllib.error.HTTPError as e:
            return e.code, e.read()

    # POST to process-combine-only
    with open(combine_path, 'rb') as f:
        files_payload = [('combine_files', os.path.basename(combine_path), f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
        status, content = post_multipart(f"{BASE_URL}/process-combine-only", files_payload)
        print('/process-combine-only ->', status)
        if status == 200:
            with open(os.path.join(OUT_DIR, 'returned_combine_processed.xlsx'), 'wb') as out:
                out.write(content)
            print('Saved returned processed combine to temp_uploads/returned_combine_processed.xlsx')
        else:
            print('Error response:', content.decode('utf-8', errors='ignore'))

    # POST to process-final-only using the saved combine file
    with open(os.path.join(OUT_DIR, 'returned_combine_processed.xlsx'), 'rb') as comb:
        with open(final_path, 'rb') as fin:
            files_payload = [
                ('combine_files', 'returned_combine_processed.xlsx', comb.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('final_mis', os.path.basename(final_path), fin.read(), 'application/vnd.openxmlformats-officedocument-spreadsheetml.sheet')
            ]
            status2, content2 = post_multipart(f"{BASE_URL}/process-final-only", files_payload)
            print('/process-final-only ->', status2)
            if status2 == 200:
                with open(os.path.join(OUT_DIR, 'returned_final_processed.xlsx'), 'wb') as out:
                    out.write(content2)
                print('Saved returned final to temp_uploads/returned_final_processed.xlsx')
            else:
                print('Error response:', content2.decode('utf-8', errors='ignore'))
